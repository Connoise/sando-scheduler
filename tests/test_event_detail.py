"""Tests for the v0.4 event-detail view and Cancel/Delete actions."""

from __future__ import annotations

import csv
import json
from datetime import date, datetime, time, timedelta
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from sando_web.app import create_app
from sando_web.config import HST, Settings
from sando_web.schedule_io import WeeklySheetCache, load_event_detail
from sando_web.schedule_write import (
    _annotate_active_match,
    _strip_status_suffix,
    update_event_status,
    write_event,
)


def _settings(tmp_path: Path) -> Settings:
    sd = tmp_path / "Schedule"
    sd.mkdir()
    return Settings(
        schedule_dir=sd,
        reminders_file=sd / "reminders.json",
        changelog_file=sd / "changelog.csv",
        config_file=sd / "config.json",
        archive_dir=sd / "Archive",
        defaults={
            "default_event_duration_minutes": 60,
            "default_reminder_offsets": ["day_of", "1h_before"],
            "morning_time": "8:00 AM",
        },
    )


def _add_dentist(settings: Settings) -> None:
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    write_event(
        event_name="Dentist",
        start=start,
        end=start + timedelta(hours=1),
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
        reminder_specs=[],
    )


# --- Pure helpers --------------------------------------------------------

def test_strip_status_suffix_recognises_both():
    assert _strip_status_suffix("Yoga (cancelled)") == ("Yoga", "cancelled")
    assert _strip_status_suffix("Yoga (deleted)") == ("Yoga", "deleted")
    assert _strip_status_suffix("Yoga") == ("Yoga", None)


def test_annotate_active_match_only_touches_named_active_entries():
    text = "Dentist | Call Mom | Dentist (cancelled)"
    new, changed = _annotate_active_match(text, "Dentist", "(cancelled)")
    assert changed is True
    assert new == "Dentist (cancelled) | Call Mom | Dentist (cancelled)"


def test_annotate_active_match_returns_unchanged_when_no_match():
    new, changed = _annotate_active_match("Yoga", "Dentist", "(cancelled)")
    assert changed is False
    assert new == "Yoga"


# --- load_event_detail ---------------------------------------------------

def test_load_event_detail_aggregates_consecutive_slots(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    detail = load_event_detail(
        "Dentist",
        date(2026, 5, 15),
        settings.schedule_dir,
        settings.reminders_file,
        WeeklySheetCache(),
    )
    assert detail is not None
    assert detail.name == "Dentist"
    assert detail.start == time(14, 0)
    assert detail.end == time(15, 0)
    assert detail.status == "active"
    assert detail.slot_count == 2


def test_load_event_detail_returns_none_when_missing(tmp_path: Path):
    settings = _settings(tmp_path)
    detail = load_event_detail(
        "Nothing",
        date(2026, 5, 15),
        settings.schedule_dir,
        settings.reminders_file,
        WeeklySheetCache(),
    )
    assert detail is None


def test_load_event_detail_picks_severest_status(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    update_event_status(
        event_name="Dentist",
        event_date=date(2026, 5, 15),
        new_status="cancelled",
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
    )
    detail = load_event_detail(
        "Dentist",
        date(2026, 5, 15),
        settings.schedule_dir,
        settings.reminders_file,
        WeeklySheetCache(),
    )
    assert detail.status == "cancelled"


# --- update_event_status -------------------------------------------------

def test_update_event_status_cancel_rewrites_cells_and_reminders(tmp_path: Path):
    settings = _settings(tmp_path)
    # Add Dentist with default reminders.
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    write_event(
        event_name="Dentist",
        start=start,
        end=start + timedelta(hours=1),
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
        reminder_specs=[],
    )
    # Add an extra reminder via direct JSON edit so we can check both.
    data = json.loads(settings.reminders_file.read_text() or "[]")
    data.append({
        "id": "stub_dentist",
        "event_name": "Dentist",
        "event_date": "2026-05-15",
        "event_start": "2:00 PM",
        "event_end": "3:00 PM",
        "remind_at": "2026-05-15T13:00:00-10:00",
        "sent": False,
        "status": "active",
        "message": "manual",
    })
    settings.reminders_file.write_text(json.dumps(data))

    count = update_event_status(
        event_name="Dentist",
        event_date=date(2026, 5, 15),
        new_status="cancelled",
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
    )
    assert count == 2  # 2:00 PM and 2:30 PM cells

    wb = load_workbook(str(settings.schedule_dir / "2026" / "2026_May.xlsx"))
    ws = wb["May 10-16"]
    assert ws.cell(row=30, column=7).value == "Dentist (cancelled)"
    assert ws.cell(row=31, column=7).value == "Dentist (cancelled)"

    reminders = json.loads(settings.reminders_file.read_text())
    statuses = [r["status"] for r in reminders if r["event_name"] == "Dentist"]
    assert statuses and all(s == "cancelled" for s in statuses)

    rows = list(csv.reader(settings.changelog_file.open()))
    actions = [r[1] for r in rows[1:]]
    assert "CANCEL" in actions


def test_update_event_status_delete_uses_deleted_suffix(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    update_event_status(
        event_name="Dentist",
        event_date=date(2026, 5, 15),
        new_status="deleted",
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
    )
    wb = load_workbook(str(settings.schedule_dir / "2026" / "2026_May.xlsx"))
    ws = wb["May 10-16"]
    assert ws.cell(row=30, column=7).value == "Dentist (deleted)"
    rows = list(csv.reader(settings.changelog_file.open()))
    assert any(r[1] == "DELETE" for r in rows[1:])


def test_update_event_status_skips_already_annotated(tmp_path: Path):
    """Re-cancelling a cancelled event is a no-op on the spreadsheet."""
    settings = _settings(tmp_path)
    _add_dentist(settings)
    update_event_status(
        event_name="Dentist",
        event_date=date(2026, 5, 15),
        new_status="cancelled",
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
    )
    again = update_event_status(
        event_name="Dentist",
        event_date=date(2026, 5, 15),
        new_status="cancelled",
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
    )
    assert again == 0


def test_update_event_status_preserves_other_cell_entries(tmp_path: Path):
    settings = _settings(tmp_path)
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    write_event(
        event_name="Dentist", start=start, end=start + timedelta(minutes=30),
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
        reminder_specs=[],
    )
    write_event(
        event_name="Call Mom", start=start, end=start + timedelta(minutes=30),
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
        reminder_specs=[],
    )
    update_event_status(
        event_name="Dentist",
        event_date=date(2026, 5, 15),
        new_status="cancelled",
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
    )
    wb = load_workbook(str(settings.schedule_dir / "2026" / "2026_May.xlsx"))
    cell = wb["May 10-16"].cell(row=30, column=7).value
    assert cell == "Dentist (cancelled) | Call Mom"


def test_update_event_status_returns_zero_when_workbook_missing(tmp_path: Path):
    settings = _settings(tmp_path)
    n = update_event_status(
        event_name="Nope",
        event_date=date(2026, 5, 15),
        new_status="cancelled",
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
    )
    assert n == 0


def test_update_event_status_rejects_bad_status(tmp_path: Path):
    settings = _settings(tmp_path)
    with pytest.raises(ValueError):
        update_event_status(
            event_name="X",
            event_date=date(2026, 5, 15),
            new_status="bogus",
            schedule_dir=settings.schedule_dir,
            reminders_file=settings.reminders_file,
            changelog_file=settings.changelog_file,
        )


# --- Routes --------------------------------------------------------------

def test_event_detail_route_renders(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    client = TestClient(create_app(settings))
    resp = client.get("/event/2026-05-15/Dentist")
    assert resp.status_code == 200
    assert "Dentist" in resp.text
    assert "2:00 PM" in resp.text and "3:00 PM" in resp.text
    assert "Cancel event" in resp.text
    assert "Delete event" in resp.text


def test_event_detail_route_404_when_missing(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    resp = client.get("/event/2026-05-15/Nothing")
    assert resp.status_code == 404


def test_event_detail_route_rejects_bad_date(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    resp = client.get("/event/not-a-date/Dentist")
    assert resp.status_code == 400


def test_post_cancel_redirects_and_updates(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    client = TestClient(create_app(settings), follow_redirects=False)
    resp = client.post("/event/2026-05-15/Dentist/cancel")
    assert resp.status_code == 303
    assert resp.headers["location"] == "/week/2026-05-15?flash=cancelled"
    wb = load_workbook(str(settings.schedule_dir / "2026" / "2026_May.xlsx"))
    assert wb["May 10-16"].cell(row=30, column=7).value == "Dentist (cancelled)"


def test_post_delete_redirects_and_updates(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    client = TestClient(create_app(settings), follow_redirects=False)
    resp = client.post("/event/2026-05-15/Dentist/delete")
    assert resp.status_code == 303
    assert resp.headers["location"] == "/week/2026-05-15?flash=deleted"


def test_post_action_404_when_event_missing(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)), follow_redirects=False)
    resp = client.post("/event/2026-05-15/Nothing/cancel")
    assert resp.status_code == 404


def test_post_unknown_action_404(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    client = TestClient(create_app(settings), follow_redirects=False)
    resp = client.post("/event/2026-05-15/Dentist/explode")
    assert resp.status_code == 404


def test_week_view_links_to_detail(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    client = TestClient(create_app(settings))
    resp = client.get("/week/2026-05-15")
    assert resp.status_code == 200
    assert 'href="/event/2026-05-15/Dentist?back=/week/2026-05-15"' in resp.text


def test_detail_view_hides_actions_for_non_active(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    update_event_status(
        event_name="Dentist",
        event_date=date(2026, 5, 15),
        new_status="cancelled",
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
    )
    client = TestClient(create_app(settings))
    resp = client.get("/event/2026-05-15/Dentist")
    assert resp.status_code == 200
    assert "Cancel event" not in resp.text
    assert "Delete event" not in resp.text


def test_cache_busts_after_cancel(tmp_path: Path):
    settings = _settings(tmp_path)
    _add_dentist(settings)
    client = TestClient(create_app(settings), follow_redirects=False)
    client.get("/week/2026-05-15")  # prime cache
    client.post("/event/2026-05-15/Dentist/cancel")
    resp = client.get("/week/2026-05-15")
    assert "Dentist" in resp.text
    # The cell now reads "Dentist (cancelled)" — the renderer splits it
    # so the chip shows "Dentist" with status-cancelled styling.
    assert "status-cancelled" in resp.text
