"""Tests for the v0.3 add-event write path.

Covers the spreadsheet write, reminders.json append, and changelog.csv
append; plus the helpers that compute covered slots and reminder timing.
"""

from __future__ import annotations

import csv
import json
from datetime import date, datetime, time, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook

from sando_web.config import HST, Settings
from sando_web.schedule_write import (
    ReminderSpec,
    _append_to_cell,
    _covered_slots,
    _slot_row,
    build_default_reminder_specs,
    build_reminder_spec,
    write_event,
)


def _settings(tmp_path: Path) -> Settings:
    sd = tmp_path / "Schedule"
    sd.mkdir()
    return Settings(
        schedule_dir=sd,
        reminders_file=sd / "reminders.json",
        changelog_file=sd / "changelog.csv",
        config_file=sd / "config.json",
        archive_dir=sd / "Archive",
        defaults={
            "default_event_duration_minutes": 60,
            "default_reminder_offsets": ["day_of", "1h_before"],
            "morning_time": "8:00 AM",
        },
    )


def test_slot_row_maps_to_spreadsheet_layout():
    # 12:00 AM is row 2 (header above), 8:00 AM is row 18, 11:30 PM is row 49.
    assert _slot_row(time(0, 0)) == 2
    assert _slot_row(time(8, 0)) == 18
    assert _slot_row(time(23, 30)) == 49


def test_covered_slots_basic_one_hour():
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    end = start + timedelta(hours=1)
    cells = _covered_slots(start, end)
    assert cells == [
        (date(2026, 5, 15), time(14, 0)),
        (date(2026, 5, 15), time(14, 30)),
    ]


def test_covered_slots_floors_non_aligned_start():
    # 14:10 → starts in the 2:00 PM cell.
    start = datetime(2026, 5, 15, 14, 10, tzinfo=HST)
    end = datetime(2026, 5, 15, 15, 0, tzinfo=HST)
    cells = _covered_slots(start, end)
    assert cells[0] == (date(2026, 5, 15), time(14, 0))
    assert (date(2026, 5, 15), time(14, 30)) in cells
    assert (date(2026, 5, 15), time(15, 0)) not in cells  # boundary excluded


def test_covered_slots_crosses_midnight():
    start = datetime(2026, 5, 15, 23, 30, tzinfo=HST)
    end = datetime(2026, 5, 16, 0, 30, tzinfo=HST)
    cells = _covered_slots(start, end)
    assert cells == [
        (date(2026, 5, 15), time(23, 30)),
        (date(2026, 5, 16), time(0, 0)),
    ]


def test_append_to_cell_joins_with_pipe():
    assert _append_to_cell(None, "Yoga") == "Yoga"
    assert _append_to_cell("", "Yoga") == "Yoga"
    assert _append_to_cell("Dentist", "Call Mom") == "Dentist | Call Mom"


def test_build_default_reminder_specs_uses_day_of_and_one_hour_before():
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    specs = build_default_reminder_specs(
        start, {"default_reminder_offsets": ["day_of", "1h_before"],
                "morning_time": "8:00 AM"}
    )
    assert [s.label for s in specs] == ["Day of", "1 hour before"]
    assert specs[0].remind_at == datetime(2026, 5, 15, 8, 0, tzinfo=HST)
    assert specs[1].remind_at == datetime(2026, 5, 15, 13, 0, tzinfo=HST)


def test_build_reminder_spec_night_before_and_30m():
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    night = build_reminder_spec("night_before", start)
    assert night.remind_at == datetime(2026, 5, 14, 20, 0, tzinfo=HST)
    half = build_reminder_spec("30m_before", start)
    assert half.remind_at == datetime(2026, 5, 15, 13, 30, tzinfo=HST)


def test_build_reminder_spec_custom_attaches_tz():
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    naive = datetime(2026, 5, 15, 10, 0)
    spec = build_reminder_spec("custom", start, custom=naive)
    assert spec.remind_at == datetime(2026, 5, 15, 10, 0, tzinfo=HST)


# --- Full write path -----------------------------------------------------

def test_write_event_creates_workbook_sheet_cells_reminders_and_log(tmp_path: Path):
    settings = _settings(tmp_path)
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    end = start + timedelta(hours=1)
    specs = build_default_reminder_specs(start, settings.defaults)

    result = write_event(
        event_name="Dentist Appointment",
        start=start,
        end=end,
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
        reminder_specs=specs,
    )

    # Workbook lives at <schedule_dir>/2026/2026_May.xlsx with sheet 'May 10-16'.
    wb_path = settings.schedule_dir / "2026" / "2026_May.xlsx"
    assert result.workbook_path == wb_path
    assert wb_path.exists()
    wb = load_workbook(str(wb_path))
    assert "May 10-16" in wb.sheetnames
    ws = wb["May 10-16"]
    # Friday is column 7 (Sun=2..Sat=8). 2:00 PM is row 30, 2:30 PM is row 31.
    assert ws.cell(row=30, column=7).value == "Dentist Appointment"
    assert ws.cell(row=31, column=7).value == "Dentist Appointment"
    # Header and time column populated.
    assert ws.cell(row=1, column=7).value.startswith("Friday")
    assert ws.cell(row=18, column=1).value == "8:00 AM"

    # Reminders appended with the right schema.
    reminders = json.loads(settings.reminders_file.read_text())
    assert len(reminders) == 2
    by_label = {r["message"]: r for r in reminders}
    assert all(r["event_name"] == "Dentist Appointment" for r in reminders)
    assert all(r["event_date"] == "2026-05-15" for r in reminders)
    assert all(r["event_start"] == "2:00 PM" for r in reminders)
    assert all(r["event_end"] == "3:00 PM" for r in reminders)
    assert all(r["status"] == "active" and r["sent"] is False for r in reminders)
    # IDs follow the spec convention: <YYYYMMDD>_<HHMMSS>_<slug>.
    assert all(r["id"].startswith("20260515_140000_dentist_appointment")
               for r in reminders)

    # Changelog written with header.
    rows = list(csv.reader(settings.changelog_file.open()))
    assert rows[0] == [
        "timestamp", "action", "event_name", "date",
        "time_start", "time_end", "details",
    ]
    assert rows[1][1:7] == [
        "ADD", "Dentist Appointment", "2026-05-15", "2:00 PM", "3:00 PM",
        "Added via mobile app",
    ]


def test_write_event_appends_to_existing_cell(tmp_path: Path):
    settings = _settings(tmp_path)
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    write_event(
        event_name="Dentist",
        start=start,
        end=start + timedelta(minutes=30),
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
        reminder_specs=[],
    )
    write_event(
        event_name="Call Mom",
        start=start,
        end=start + timedelta(minutes=30),
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
        reminder_specs=[],
    )
    wb = load_workbook(str(settings.schedule_dir / "2026" / "2026_May.xlsx"))
    ws = wb["May 10-16"]
    assert ws.cell(row=30, column=7).value == "Dentist | Call Mom"


def test_write_event_preserves_existing_reminders(tmp_path: Path):
    settings = _settings(tmp_path)
    settings.reminders_file.write_text(json.dumps([
        {"id": "existing_1", "event_name": "X", "status": "active"}
    ]))
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    specs = build_default_reminder_specs(start, settings.defaults)
    write_event(
        event_name="Y",
        start=start,
        end=start + timedelta(hours=1),
        schedule_dir=settings.schedule_dir,
        reminders_file=settings.reminders_file,
        changelog_file=settings.changelog_file,
        reminder_specs=specs,
    )
    reminders = json.loads(settings.reminders_file.read_text())
    ids = [r["id"] for r in reminders]
    assert "existing_1" in ids
    assert len(reminders) == 3


def test_write_event_rejects_naive_datetime(tmp_path: Path):
    settings = _settings(tmp_path)
    with pytest.raises(ValueError):
        write_event(
            event_name="X",
            start=datetime(2026, 5, 15, 14, 0),  # naive
            end=datetime(2026, 5, 15, 15, 0, tzinfo=HST),
            schedule_dir=settings.schedule_dir,
            reminders_file=settings.reminders_file,
            changelog_file=settings.changelog_file,
            reminder_specs=[],
        )


def test_changelog_appends_without_duplicate_header(tmp_path: Path):
    settings = _settings(tmp_path)
    start = datetime(2026, 5, 15, 14, 0, tzinfo=HST)
    for name in ("A", "B"):
        write_event(
            event_name=name,
            start=start,
            end=start + timedelta(minutes=30),
            schedule_dir=settings.schedule_dir,
            reminders_file=settings.reminders_file,
            changelog_file=settings.changelog_file,
            reminder_specs=[],
        )
    rows = list(csv.reader(settings.changelog_file.open()))
    assert rows[0][0] == "timestamp"  # only one header
    assert sum(1 for r in rows if r and r[0] == "timestamp") == 1
    assert len(rows) == 3  # header + 2 events
