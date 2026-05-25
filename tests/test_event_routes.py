"""End-to-end tests for /event/new, POST /event, and /api/events."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from sando_web.app import create_app
from sando_web.config import Settings


def _settings(tmp_path: Path) -> Settings:
    sd = tmp_path / "Schedule"
    sd.mkdir()
    return Settings(
        schedule_dir=sd,
        reminders_file=sd / "reminders.json",
        changelog_file=sd / "changelog.csv",
        config_file=sd / "config.json",
        archive_dir=sd / "Archive",
        defaults={
            "default_event_duration_minutes": 60,
            "default_reminder_offsets": ["day_of", "1h_before"],
            "morning_time": "8:00 AM",
        },
    )


def test_event_new_form_renders(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    resp = client.get("/event/new")
    assert resp.status_code == 200
    assert "Event name" in resp.text
    assert "Reminders" in resp.text
    # Default reminder rows for the two defaults.
    assert resp.text.count('<div class="reminder-row">') == 2


def test_event_new_accepts_start_param(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    resp = client.get("/event/new?start=2026-05-15")
    assert resp.status_code == 200
    assert 'value="2026-05-15T09:00"' in resp.text


def test_post_event_writes_and_redirects(tmp_path: Path):
    settings = _settings(tmp_path)
    client = TestClient(create_app(settings), follow_redirects=False)
    resp = client.post(
        "/event",
        data={
            "event_name": "Dentist",
            "start": "2026-05-15T14:00",
            "duration_minutes": "60",
            "reminder": ["day_of", "1h_before"],
        },
    )
    assert resp.status_code == 303
    assert resp.headers["location"] == "/week/2026-05-15?flash=added"

    wb = load_workbook(str(settings.schedule_dir / "2026" / "2026_May.xlsx"))
    ws = wb["May 10-16"]
    assert ws.cell(row=30, column=7).value == "Dentist"
    reminders = json.loads(settings.reminders_file.read_text())
    assert len(reminders) == 2


def test_post_event_uses_default_duration_when_omitted(tmp_path: Path):
    settings = _settings(tmp_path)
    client = TestClient(create_app(settings), follow_redirects=False)
    resp = client.post(
        "/event",
        data={
            "event_name": "Lunch",
            "start": "2026-05-15T12:00",
            "duration_minutes": "",
        },
    )
    assert resp.status_code == 303
    # 60-minute default → covers 12:00 and 12:30 cells (rows 26, 27).
    wb = load_workbook(str(settings.schedule_dir / "2026" / "2026_May.xlsx"))
    ws = wb["May 10-16"]
    assert ws.cell(row=26, column=7).value == "Lunch"
    assert ws.cell(row=27, column=7).value == "Lunch"


def test_post_event_uses_defaults_when_no_reminders_submitted(tmp_path: Path):
    settings = _settings(tmp_path)
    client = TestClient(create_app(settings), follow_redirects=False)
    client.post(
        "/event",
        data={
            "event_name": "Lunch",
            "start": "2026-05-15T12:00",
            "duration_minutes": "60",
        },
    )
    reminders = json.loads(settings.reminders_file.read_text())
    assert len(reminders) == 2  # day_of + 1h_before defaults
    labels = sorted(r["remind_at"][:16] for r in reminders)
    assert "2026-05-15T08:00" in labels[0]  # day-of @ 8 AM
    assert "2026-05-15T11:00" in labels[1]  # 1h before noon


def test_post_event_rejects_empty_name(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    resp = client.post(
        "/event",
        data={
            "event_name": "",
            "start": "2026-05-15T14:00",
            "duration_minutes": "60",
        },
    )
    assert resp.status_code == 400
    assert "required" in resp.text.lower()


def test_post_event_rejects_bad_datetime(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    resp = client.post(
        "/event",
        data={
            "event_name": "Foo",
            "start": "not-a-date",
            "duration_minutes": "60",
        },
    )
    assert resp.status_code == 400


def test_post_event_with_custom_reminder(tmp_path: Path):
    settings = _settings(tmp_path)
    client = TestClient(create_app(settings), follow_redirects=False)
    resp = client.post(
        "/event",
        data={
            "event_name": "Yoga",
            "start": "2026-05-15T18:00",
            "duration_minutes": "60",
            # Lists in dict values become repeated form fields.
            "reminder": ["custom", "day_of"],
            # Same number of reminder_custom values as reminder values;
            # the day_of slot gets an empty placeholder.
            "reminder_custom": ["2026-05-15T15:30", ""],
        },
    )
    assert resp.status_code == 303
    reminders = json.loads(settings.reminders_file.read_text())
    assert len(reminders) == 2
    remind_times = sorted(r["remind_at"][:16] for r in reminders)
    assert remind_times[0].startswith("2026-05-15T08:00")  # day_of
    assert remind_times[1].startswith("2026-05-15T15:30")  # custom


def test_api_events_returns_range(tmp_path: Path):
    settings = _settings(tmp_path)
    client = TestClient(create_app(settings), follow_redirects=False)
    client.post(
        "/event",
        data={
            "event_name": "Dentist",
            "start": "2026-05-15T14:00",
            "duration_minutes": "60",
        },
    )
    resp = client.get("/api/events?from=2026-05-15&to=2026-05-15")
    assert resp.status_code == 200
    payload = resp.json()
    names = [e["name"] for e in payload]
    assert names.count("Dentist") == 2  # 2:00 and 2:30 cells
    assert all(e["status"] == "active" for e in payload)


def test_api_events_validates_inputs(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    assert client.get("/api/events").status_code == 400
    assert client.get("/api/events?from=bad&to=2026-05-15").status_code == 400


def test_week_view_shows_flash_after_add(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    resp = client.get("/week/2026-05-15?flash=added")
    assert resp.status_code == 200
    assert "Event added" in resp.text


def test_week_view_has_add_link(tmp_path: Path):
    client = TestClient(create_app(_settings(tmp_path)))
    resp = client.get("/week/2026-05-15")
    assert 'href="/event/new"' in resp.text


def test_cache_busts_after_post(tmp_path: Path):
    """A read after an add must reflect the new event, not stale cache."""
    settings = _settings(tmp_path)
    client = TestClient(create_app(settings), follow_redirects=False)
    # Prime the cache by viewing the week first.
    client.get("/week/2026-05-15")
    client.post(
        "/event",
        data={
            "event_name": "Dentist",
            "start": "2026-05-15T14:00",
            "duration_minutes": "60",
        },
    )
    resp = client.get("/week/2026-05-15")
    assert "Dentist" in resp.text
