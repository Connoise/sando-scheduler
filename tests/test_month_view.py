from datetime import date
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from sando_web.app import create_app
from sando_web.calendar_utils import (
    half_hour_slots,
    sheet_name_for_week,
    slot_label,
    week_start,
    workbook_for_week,
)
from sando_web.config import Settings
from sando_web.schedule_io import (
    MonthView,
    WeeklySheetCache,
    load_events_in_range,
)


def _settings(tmp_path: Path) -> Settings:
    schedule_dir = tmp_path / "Schedule"
    schedule_dir.mkdir()
    return Settings(
        schedule_dir=schedule_dir,
        reminders_file=schedule_dir / "reminders.json",
        changelog_file=schedule_dir / "changelog.csv",
        config_file=schedule_dir / "config.json",
        archive_dir=schedule_dir / "Archive",
        defaults={},
    )


def _write_week_sheet(
    settings: Settings, anchor: date, cells: dict[tuple[date, str], str]
) -> Path:
    """Write a workbook containing the week that anchors `anchor`.

    `cells` maps (event_date, slot_label) → cell text.
    """
    ws_start = week_start(anchor)
    year, month_name = workbook_for_week(ws_start)
    sheet = sheet_name_for_week(ws_start)
    folder = settings.schedule_dir / str(year)
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{year}_{month_name}.xlsx"
    if path.exists():
        wb = load_workbook(str(path))
        if sheet in wb.sheetnames:
            ws = wb[sheet]
        else:
            ws = wb.create_sheet(title=sheet)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet

    days = [ws_start.fromordinal(ws_start.toordinal() + i) for i in range(7)]
    weekday_names = ["Sunday", "Monday", "Tuesday", "Wednesday",
                     "Thursday", "Friday", "Saturday"]
    for i, d in enumerate(days):
        ws.cell(
            row=1,
            column=i + 2,
            value=f"{weekday_names[i]} {d.month}/{d.day}",
        )

    label_to_row = {}
    for row_idx, slot in enumerate(half_hour_slots(), start=2):
        lbl = slot_label(slot)
        ws.cell(row=row_idx, column=1, value=lbl)
        label_to_row[lbl] = row_idx

    for (ev_date, lbl), text in cells.items():
        col = days.index(ev_date) + 2
        ws.cell(row=label_to_row[lbl], column=col, value=text)

    wb.save(str(path))
    return path


def test_load_events_in_range_crosses_weeks(tmp_path: Path):
    settings = _settings(tmp_path)
    # Two events on different weeks of May 2026.
    _write_week_sheet(
        settings, date(2026, 5, 12),
        {(date(2026, 5, 12), "9:00 AM"): "Standup"},
    )
    _write_week_sheet(
        settings, date(2026, 5, 19),
        {(date(2026, 5, 19), "2:00 PM"): "Dentist"},
    )
    cache = WeeklySheetCache()
    events = load_events_in_range(
        date(2026, 5, 1), date(2026, 5, 31), settings.schedule_dir, cache
    )
    names = sorted(e.name for e in events)
    assert names == ["Dentist", "Standup"]


def test_month_view_lays_out_full_grid(tmp_path: Path):
    settings = _settings(tmp_path)
    _write_week_sheet(
        settings, date(2026, 5, 12),
        {
            (date(2026, 5, 12), "9:00 AM"): "Standup",
            (date(2026, 5, 12), "2:00 PM"): "Lunch",
            (date(2026, 5, 12), "4:00 PM"): "Yoga",
        },
    )
    view = MonthView.build(
        2026, 5, settings.schedule_dir, WeeklySheetCache(),
        today=date(2026, 5, 15),
    )
    assert view.month_name == "May"
    # Grid spans complete weeks: starts on a Sunday, ends on a Saturday.
    assert view.grid_start.weekday() == 6  # Sunday
    assert view.grid_end.weekday() == 5    # Saturday
    assert view.grid_start <= date(2026, 5, 1)
    assert view.grid_end >= date(2026, 5, 31)
    # Each row is exactly 7 cells.
    assert all(len(row) == 7 for row in view.weeks)
    # Find May 12 in the grid.
    target = next(
        cell for row in view.weeks for cell in row if cell.date == date(2026, 5, 12)
    )
    assert target.in_month is True
    assert [e.name for e in target.events] == ["Standup", "Lunch", "Yoga"]
    assert target.overflow == 1  # display_events caps at 2
    assert len(target.display_events) == 2


def test_month_view_marks_today_and_out_of_month(tmp_path: Path):
    settings = _settings(tmp_path)
    view = MonthView.build(
        2026, 5, settings.schedule_dir, WeeklySheetCache(),
        today=date(2026, 5, 15),
    )
    today_cell = next(
        c for row in view.weeks for c in row if c.is_today
    )
    assert today_cell.date == date(2026, 5, 15)
    # The first row almost certainly contains April days (out-of-month).
    oom_count = sum(1 for row in view.weeks for c in row if not c.in_month)
    assert oom_count > 0


def test_month_view_navigation_wraps_year_boundary():
    view = MonthView.build(
        2026, 1, Path("/nonexistent"), WeeklySheetCache(),
        today=date(2026, 1, 1),
    )
    assert view.prev_month == (2025, 12)
    assert view.next_month == (2026, 2)

    view = MonthView.build(
        2026, 12, Path("/nonexistent"), WeeklySheetCache(),
        today=date(2026, 12, 1),
    )
    assert view.prev_month == (2026, 11)
    assert view.next_month == (2027, 1)


def test_month_route_renders(tmp_path: Path):
    settings = _settings(tmp_path)
    _write_week_sheet(
        settings, date(2026, 5, 12),
        {(date(2026, 5, 12), "9:00 AM"): "Standup"},
    )
    app = create_app(settings)
    client = TestClient(app)
    resp = client.get("/month/2026-05")
    assert resp.status_code == 200
    assert "May 2026" in resp.text
    assert "Standup" in resp.text
    # Day links should point to the week view.
    assert 'href="/week/2026-05-12"' in resp.text


def test_month_route_rejects_bad_input(tmp_path: Path):
    app = create_app(_settings(tmp_path))
    client = TestClient(app)
    assert client.get("/month/2026-13").status_code == 400
    assert client.get("/month/not-a-month").status_code == 400


def test_week_view_links_to_month(tmp_path: Path):
    app = create_app(_settings(tmp_path))
    client = TestClient(app)
    resp = client.get("/week/2026-05-15")
    assert resp.status_code == 200
    assert 'href="/month/2026-05"' in resp.text
