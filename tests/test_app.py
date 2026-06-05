from datetime import date, datetime, timezone
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from sando_web.app import create_app
from sando_web.calendar_utils import half_hour_slots, slot_label
from sando_web.config import Settings


class _FrozenDateTime(datetime):
    """A datetime whose `now()` is pinned to an instant where UTC and HST
    disagree: 2026-06-01 05:00 UTC is 2026-05-31 19:00 HST, so the server
    (UTC) and Hawaii fall on different days *and* different months.
    """

    _UTC_NOW = datetime(2026, 6, 1, 5, 0, tzinfo=timezone.utc)

    @classmethod
    def now(cls, tz=None):
        if tz is None:
            return cls._UTC_NOW.replace(tzinfo=None)
        return cls._UTC_NOW.astimezone(tz)


def _freeze_hst_clock(monkeypatch):
    """Pin `today_hst()` to the divergent instant defined above."""
    monkeypatch.setattr("sando_web.config.datetime", _FrozenDateTime)


def _make_settings(tmp_path: Path) -> Settings:
    schedule_dir = tmp_path / "Schedule"
    schedule_dir.mkdir()
    return Settings(
        schedule_dir=schedule_dir,
        reminders_file=schedule_dir / "reminders.json",
        changelog_file=schedule_dir / "changelog.csv",
        config_file=schedule_dir / "config.json",
        archive_dir=schedule_dir / "Archive",
        defaults={},
    )


def _write_workbook_with_event(
    settings: Settings, year: int, month_name: str, sheet: str
) -> Path:
    folder = settings.schedule_dir / str(year)
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{year}_{month_name}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    days = ["Sunday 5/10", "Monday 5/11", "Tuesday 5/12", "Wednesday 5/13",
            "Thursday 5/14", "Friday 5/15", "Saturday 5/16"]
    for i, label in enumerate(days, start=2):
        ws.cell(row=1, column=i, value=label)
    for row_idx, slot in enumerate(half_hour_slots(), start=2):
        ws.cell(row=row_idx, column=1, value=slot_label(slot))
    ws.cell(row=30, column=7, value="Dentist Appointment")  # Fri 2:00 PM
    wb.save(str(path))
    return path


def test_root_redirects_to_current_week(tmp_path: Path):
    app = create_app(_make_settings(tmp_path))
    client = TestClient(app, follow_redirects=False)
    resp = client.get("/")
    assert resp.status_code == 307
    assert resp.headers["location"].startswith("/week/")


def test_today_hst_uses_hawaii_calendar_date(monkeypatch):
    _freeze_hst_clock(monkeypatch)
    from sando_web.config import today_hst
    # UTC has already ticked over to June 1; Hawaii is still on May 31.
    assert today_hst() == date(2026, 5, 31)


def test_root_redirects_to_week_of_hst_today(monkeypatch, tmp_path: Path):
    _freeze_hst_clock(monkeypatch)
    app = create_app(_make_settings(tmp_path))
    client = TestClient(app, follow_redirects=False)
    resp = client.get("/")
    assert resp.status_code == 307
    # Must default off the HST date (May 31), not the server/UTC date (Jun 1).
    assert resp.headers["location"] == "/week/2026-05-31"


def test_month_root_redirects_to_current_hst_month(monkeypatch, tmp_path: Path):
    _freeze_hst_clock(monkeypatch)
    app = create_app(_make_settings(tmp_path))
    client = TestClient(app, follow_redirects=False)
    resp = client.get("/month")
    assert resp.status_code == 307
    # HST month is May (05), not the server/UTC month June (06).
    assert resp.headers["location"] == "/month/2026-05"


def test_week_renders_event_from_workbook(tmp_path: Path):
    settings = _make_settings(tmp_path)
    _write_workbook_with_event(settings, 2026, "May", "May 10-16")
    app = create_app(settings)
    client = TestClient(app)
    resp = client.get("/week/2026-05-15")
    assert resp.status_code == 200
    assert "Dentist Appointment" in resp.text
    assert "May 10" in resp.text and "May 16" in resp.text


def test_week_renders_empty_when_workbook_missing(tmp_path: Path):
    app = create_app(_make_settings(tmp_path))
    client = TestClient(app)
    resp = client.get("/week/2026-05-15")
    assert resp.status_code == 200
    assert "workbook not yet created" in resp.text


def test_week_rejects_bad_date(tmp_path: Path):
    app = create_app(_make_settings(tmp_path))
    client = TestClient(app)
    resp = client.get("/week/not-a-date")
    assert resp.status_code == 400


def test_healthz(tmp_path: Path):
    app = create_app(_make_settings(tmp_path))
    client = TestClient(app)
    resp = client.get("/healthz")
    assert resp.status_code == 200
    assert resp.json()["status"] == "ok"
