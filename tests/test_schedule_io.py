import json
from datetime import date, time
from pathlib import Path

import pytest
from openpyxl import Workbook

from sando_web.calendar_utils import WeekContext, half_hour_slots, slot_label
from sando_web.schedule_io import (
    WeeklySheetCache,
    WeeklySheetView,
    _parse_cell,
    load_reminders,
)


def _write_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "May 10-16"

    # Header row: column A blank, B..H = Sun..Sat for week of 2026-05-10.
    days = ["Sunday 5/10", "Monday 5/11", "Tuesday 5/12", "Wednesday 5/13",
            "Thursday 5/14", "Friday 5/15", "Saturday 5/16"]
    for i, label in enumerate(days, start=2):
        ws.cell(row=1, column=i, value=label)

    # Slot column A: 48 half-hour labels.
    for row_idx, slot in enumerate(half_hour_slots(), start=2):
        ws.cell(row=row_idx, column=1, value=slot_label(slot))

    # 2:00 PM Friday (column G = day 5 + column offset 2 = col 7): two events.
    # 2:00 PM = index 28 (12 AM = idx 0), so row = 2 + 28 = 30.
    ws.cell(row=30, column=7, value="Dentist Appointment | Call Mom")
    # 8:00 AM Sunday: cancelled event. 8:00 AM idx = 16, row = 18, col B = 2.
    ws.cell(row=18, column=2, value="Yoga (cancelled)")
    # 9:30 AM Wednesday: deleted event. row = 21, col E = 5.
    ws.cell(row=21, column=5, value="Standup (deleted)")

    wb.save(str(path))


def test_parse_cell_splits_and_extracts_status():
    assert _parse_cell("Foo | Bar") == [("Foo", "active"), ("Bar", "active")]
    assert _parse_cell("Yoga (cancelled)") == [("Yoga", "cancelled")]
    assert _parse_cell("Standup (deleted) | Notes") == [
        ("Standup", "deleted"),
        ("Notes", "active"),
    ]
    assert _parse_cell("") == []
    assert _parse_cell(None) == []
    assert _parse_cell("   ") == []


def test_weekly_view_parses_real_workbook(tmp_path: Path):
    wb_path = tmp_path / "2026_May.xlsx"
    _write_sample_workbook(wb_path)

    ctx = WeekContext.for_date(date(2026, 5, 15), today=date(2026, 5, 15))
    assert ctx.sheet_name == "May 10-16"

    cache = WeeklySheetCache()
    view = WeeklySheetView.build(ctx, wb_path, cache)
    assert view.workbook_exists is True

    # Pull cells at the slots we wrote.
    by_label = {row.label: row for row in view.rows}

    fri_idx = 5  # Sunday=0..Saturday=6
    sun_idx = 0
    wed_idx = 3

    fri_2pm = by_label["2:00 PM"].cells[fri_idx]
    assert [e.name for e in fri_2pm] == ["Dentist Appointment", "Call Mom"]
    assert all(e.status == "active" for e in fri_2pm)
    assert fri_2pm[0].date == date(2026, 5, 15)

    sun_8am = by_label["8:00 AM"].cells[sun_idx]
    assert len(sun_8am) == 1 and sun_8am[0].name == "Yoga"
    assert sun_8am[0].status == "cancelled"

    wed_930 = by_label["9:30 AM"].cells[wed_idx]
    assert len(wed_930) == 1 and wed_930[0].name == "Standup"
    assert wed_930[0].status == "deleted"


def test_weekly_view_missing_workbook(tmp_path: Path):
    ctx = WeekContext.for_date(date(2026, 5, 15))
    view = WeeklySheetView.build(ctx, tmp_path / "nope.xlsx", WeeklySheetCache())
    assert view.workbook_exists is False
    # Should still build 48 empty rows so the template can render.
    assert len(view.rows) == 48
    assert all(all(cell == [] for cell in row.cells) for row in view.rows)


def test_cache_invalidates_on_mtime_change(tmp_path: Path):
    import os, time as _time

    wb_path = tmp_path / "2026_May.xlsx"
    _write_sample_workbook(wb_path)
    ctx = WeekContext.for_date(date(2026, 5, 15))
    cache = WeeklySheetCache()

    v1 = WeeklySheetView.build(ctx, wb_path, cache)
    n1 = sum(len(row.cells[5]) for row in v1.rows)  # Friday events
    assert n1 == 2

    # Rewrite with no Friday event and bump mtime.
    wb = Workbook()
    ws = wb.active
    ws.title = "May 10-16"
    for i, label in enumerate(
        ["Sunday 5/10", "Monday 5/11", "Tuesday 5/12", "Wednesday 5/13",
         "Thursday 5/14", "Friday 5/15", "Saturday 5/16"], start=2):
        ws.cell(row=1, column=i, value=label)
    for row_idx, slot in enumerate(half_hour_slots(), start=2):
        ws.cell(row=row_idx, column=1, value=slot_label(slot))
    wb.save(str(wb_path))
    os.utime(wb_path, (wb_path.stat().st_atime + 5, wb_path.stat().st_mtime + 5))

    v2 = WeeklySheetView.build(ctx, wb_path, cache)
    n2 = sum(len(row.cells[5]) for row in v2.rows)
    assert n2 == 0


def test_load_reminders_round_trip(tmp_path: Path):
    rfile = tmp_path / "reminders.json"
    payload = [
        {
            "id": "abc",
            "event_name": "Yoga",
            "event_date": "2026-05-15",
            "event_start": "6:00 PM",
            "event_end": "7:00 PM",
            "remind_at": "2026-05-15T17:00:00-10:00",
            "sent": False,
            "status": "active",
            "message": "Yoga reminder",
        },
        {"id": "broken", "event_date": "not-a-date"},
        "garbage",
    ]
    rfile.write_text(json.dumps(payload))
    reminders = load_reminders(rfile)
    assert len(reminders) == 2  # "garbage" string is dropped, broken date kept
    assert reminders[0].event_name == "Yoga"
    assert reminders[0].event_date == date(2026, 5, 15)
    assert reminders[1].event_date is None


def test_load_reminders_missing_file(tmp_path: Path):
    assert load_reminders(tmp_path / "nope.json") == []
