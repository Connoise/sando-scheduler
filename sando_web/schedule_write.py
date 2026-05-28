"""Write-side I/O: create events in the xlsx, reminders.json, and changelog.csv.

Concurrency model per mobile_app_plan.md §Concurrency:
  - xlsx: sibling `.lock` file held via fcntl.flock(LOCK_EX) for the
    duration of the read-modify-write cycle (openpyxl is not safe for
    concurrent writes).
  - reminders.json: LOCK_EX on the file handle around the full
    read-modify-write.
  - changelog.csv: LOCK_EX on the file handle around the append.
"""

from __future__ import annotations

import csv
import fcntl
import json
import logging
import re
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Iterator

from openpyxl import Workbook, load_workbook

from .calendar_utils import (
    MONTH_NAMES,
    WeekContext,
    half_hour_slots,
    sheet_name_for_week,
    slot_label,
    week_dates,
    week_start,
    workbook_for_week,
)
from .config import HST


logger = logging.getLogger(__name__)


# Match the spreadsheet's "Event Name (cancelled)" / "(deleted)" suffix.
_STATUS_SUFFIX = {
    "cancelled": "(cancelled)",
    "deleted": "(deleted)",
}


# --- File locking ---------------------------------------------------------

@contextmanager
def _xlsx_lock(workbook_path: Path) -> Iterator[None]:
    """Hold an exclusive lock on a sibling `.lock` file."""
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    lock_path = workbook_path.with_suffix(workbook_path.suffix + ".lock")
    with lock_path.open("a+") as lf:
        fcntl.flock(lf, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(lf, fcntl.LOCK_UN)


# --- Workbook & sheet bootstrapping --------------------------------------

def _ensure_weekly_sheet(wb: Workbook, week_start_date: date):
    """Create the weekly sheet for `week_start_date` if missing.

    Returns the worksheet. Header row uses 'Sunday M/D' style labels (the
    same form `_parse_header_label` already understands), and column A is
    populated with all 48 half-hour labels.
    """
    sheet_name = sheet_name_for_week(week_start_date)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    # If the default 'Sheet' is still empty and untouched, repurpose it
    # instead of leaving it behind.
    if wb.sheetnames == ["Sheet"] and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        ws = wb["Sheet"]
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)

    weekday_names = ["Sunday", "Monday", "Tuesday", "Wednesday",
                     "Thursday", "Friday", "Saturday"]
    for i, d in enumerate(week_dates(week_start_date)):
        ws.cell(row=1, column=i + 2,
                value=f"{weekday_names[i]} {d.month}/{d.day}")

    for row_idx, slot in enumerate(half_hour_slots(), start=2):
        ws.cell(row=row_idx, column=1, value=slot_label(slot))

    return ws


# --- Event writing --------------------------------------------------------

def _slot_row(slot: time) -> int:
    """Row number for a half-hour slot. 12:00 AM = row 2."""
    idx = slot.hour * 2 + (1 if slot.minute >= 30 else 0)
    return 2 + idx


def _day_column(day_offset_from_sunday: int) -> int:
    """Column for a day. Sunday=0 -> col 2 ... Saturday=6 -> col 8."""
    return 2 + day_offset_from_sunday


def _half_hour_floor(t: time) -> time:
    minute = 0 if t.minute < 30 else 30
    return time(hour=t.hour, minute=minute)


def _covered_slots(start: datetime, end: datetime) -> list[tuple[date, time]]:
    """Half-hour cells covered by [start, end). Events ending exactly on a
    slot boundary do not occupy the boundary cell.

    Splits across midnight if start/end cross a day boundary.
    """
    if end <= start:
        end = start + timedelta(minutes=30)
    floor_start = datetime.combine(
        start.date(), _half_hour_floor(start.time()), tzinfo=start.tzinfo
    )
    out: list[tuple[date, time]] = []
    cur = floor_start
    while cur < end:
        out.append((cur.date(), cur.time()))
        cur += timedelta(minutes=30)
    return out


def _append_to_cell(existing: object, name: str) -> str:
    """Add an event name to a cell, joining with ' | ' if non-empty."""
    if existing is None:
        return name
    text = str(existing).strip()
    if not text:
        return name
    return f"{text} | {name}"


# --- Reminders -----------------------------------------------------------

@dataclass(frozen=True)
class ReminderSpec:
    """One reminder to register for a new event."""

    remind_at: datetime  # tz-aware
    label: str  # "Day of", "1 hour before", etc., for the changelog


def _slugify(name: str) -> str:
    """Match the id convention in scheduler_instructions.md §reminders.json."""
    s = name.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s or "event"


def _reminder_id(event_name: str, start: datetime) -> str:
    return f"{start.strftime('%Y%m%d_%H%M%S')}_{_slugify(event_name)}"


def _reminder_message(event_name: str, start_label: str, remind_at: datetime,
                      event_date: date) -> str:
    """Build a human-readable Telegram message body."""
    if remind_at.date() == event_date:
        when = "today"
    elif remind_at.date() == event_date - timedelta(days=1):
        when = "tomorrow"
    else:
        when = f"on {event_date.strftime('%a %b %-d')}"
    return f"Reminder: {event_name} {when} at {start_label}"


def build_default_reminder_specs(
    start: datetime, defaults: dict
) -> list[ReminderSpec]:
    """Mirror scheduler_instructions.md §Default Reminder Timing."""
    offsets = defaults.get("default_reminder_offsets", ["day_of", "1h_before"])
    morning_label = defaults.get("morning_time", "8:00 AM")
    morning_t = _parse_clock_label(morning_label) or time(8, 0)
    specs: list[ReminderSpec] = []
    for off in offsets:
        spec = build_reminder_spec(off, start, morning_t)
        if spec is not None:
            specs.append(spec)
    return specs


def build_reminder_spec(
    kind: str, start: datetime, morning: time = time(8, 0),
    custom: datetime | None = None,
) -> ReminderSpec | None:
    """Translate a quick-pick key into a concrete ReminderSpec."""
    if kind == "day_of":
        remind = datetime.combine(start.date(), morning, tzinfo=start.tzinfo)
        return ReminderSpec(remind_at=remind, label="Day of")
    if kind == "1h_before":
        return ReminderSpec(
            remind_at=start - timedelta(hours=1), label="1 hour before"
        )
    if kind == "30m_before":
        return ReminderSpec(
            remind_at=start - timedelta(minutes=30), label="30 min before"
        )
    if kind == "night_before":
        night = datetime.combine(
            start.date() - timedelta(days=1), time(20, 0), tzinfo=start.tzinfo
        )
        return ReminderSpec(remind_at=night, label="Night before")
    if kind == "custom" and custom is not None:
        if custom.tzinfo is None:
            custom = custom.replace(tzinfo=start.tzinfo)
        return ReminderSpec(remind_at=custom, label="Custom")
    return None


def _parse_clock_label(s: str) -> time | None:
    s = s.strip()
    for fmt in ("%I:%M %p", "%I:%M%p", "%I %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


# --- The public API -------------------------------------------------------

@dataclass
class WriteResult:
    workbook_path: Path
    sheet_name: str
    reminder_ids: list[str]


def write_event(
    *,
    event_name: str,
    start: datetime,
    end: datetime,
    schedule_dir: Path,
    reminders_file: Path,
    changelog_file: Path,
    reminder_specs: list[ReminderSpec],
    details: str = "Added via mobile app",
) -> WriteResult:
    """Dual-write the event to xlsx + reminders.json + changelog.csv.

    The start/end datetimes must be timezone-aware (HST in practice).
    """
    if start.tzinfo is None or end.tzinfo is None:
        raise ValueError("start/end must be timezone-aware")
    event_name = event_name.strip()
    if not event_name:
        raise ValueError("event_name is required")

    # 1) Write to spreadsheet (and bootstrap workbook/sheet if needed).
    cells = _covered_slots(start, end)
    workbook_path = _write_cells(schedule_dir, event_name, cells)

    # 2) Append reminder entries.
    reminder_ids = _append_reminders(
        reminders_file=reminders_file,
        event_name=event_name,
        start=start,
        end=end,
        specs=reminder_specs,
    )

    # 3) Log to changelog.
    _append_changelog(
        changelog_file,
        action="ADD",
        event_name=event_name,
        ev_date=start.date(),
        start=start,
        end=end,
        details=details,
    )

    return WriteResult(
        workbook_path=workbook_path,
        sheet_name=sheet_name_for_week(week_start(start.date())),
        reminder_ids=reminder_ids,
    )


def _write_cells(
    schedule_dir: Path,
    event_name: str,
    cells: list[tuple[date, time]],
) -> Path:
    """Write `event_name` into each (date, slot) cell.

    Groups cells by workbook (events normally fit in one week, but across
    midnight at month boundaries we may touch two workbooks).
    """
    if not cells:
        raise ValueError("event covers no cells")

    by_workbook: dict[Path, list[tuple[date, time]]] = {}
    for d, slot in cells:
        ws_start = week_start(d)
        year, month_name = workbook_for_week(ws_start)
        wb_path = schedule_dir / str(year) / f"{year}_{month_name}.xlsx"
        by_workbook.setdefault(wb_path, []).append((d, slot))

    # Touch the workbook of the event's start date first so callers get a
    # sensible "primary" path back.
    primary = schedule_dir / str(cells[0][0].year)
    primary_path: Path | None = None

    for wb_path, group in by_workbook.items():
        with _xlsx_lock(wb_path):
            if wb_path.exists():
                wb = load_workbook(str(wb_path))
            else:
                wb = Workbook()
            try:
                # Pre-create every week touched by this group.
                touched_weeks = {week_start(d) for d, _ in group}
                ws_by_week = {ws: _ensure_weekly_sheet(wb, ws)
                              for ws in touched_weeks}

                for d, slot in group:
                    ws = ws_by_week[week_start(d)]
                    row = _slot_row(slot)
                    col = _day_column((d - week_start(d)).days)
                    existing = ws.cell(row=row, column=col).value
                    ws.cell(row=row, column=col,
                            value=_append_to_cell(existing, event_name))

                wb.save(str(wb_path))
            finally:
                wb.close()
        if wb_path.parent == primary and primary_path is None:
            primary_path = wb_path
    return primary_path or next(iter(by_workbook))


def _append_reminders(
    *,
    reminders_file: Path,
    event_name: str,
    start: datetime,
    end: datetime,
    specs: list[ReminderSpec],
) -> list[str]:
    """Atomically append reminder entries with an exclusive lock."""
    reminders_file.parent.mkdir(parents=True, exist_ok=True)
    if not reminders_file.exists():
        reminders_file.write_text("[]")

    new_ids: list[str] = []
    base_id = _reminder_id(event_name, start)
    start_label = _format_clock(start.time())
    end_label = _format_clock(end.time())

    with reminders_file.open("r+") as f:
        fcntl.flock(f, fcntl.LOCK_EX)
        try:
            try:
                data = json.load(f)
                if not isinstance(data, list):
                    data = []
            except json.JSONDecodeError:
                data = []

            for i, spec in enumerate(specs):
                entry_id = base_id if len(specs) == 1 else f"{base_id}_{i}"
                new_ids.append(entry_id)
                msg = _reminder_message(
                    event_name, start_label, spec.remind_at, start.date()
                )
                data.append({
                    "id": entry_id,
                    "event_name": event_name,
                    "event_date": start.date().isoformat(),
                    "event_start": start_label,
                    "event_end": end_label,
                    "remind_at": spec.remind_at.isoformat(),
                    "sent": False,
                    "status": "active",
                    "message": msg,
                })

            f.seek(0)
            f.truncate()
            json.dump(data, f, indent=2)
        finally:
            fcntl.flock(f, fcntl.LOCK_UN)

    return new_ids


CHANGELOG_HEADER = [
    "timestamp", "action", "event_name", "date",
    "time_start", "time_end", "details",
]


def _append_changelog(
    changelog_file: Path,
    *,
    action: str,
    event_name: str,
    ev_date: date,
    start: datetime,
    end: datetime,
    details: str,
) -> None:
    changelog_file.parent.mkdir(parents=True, exist_ok=True)
    needs_header = not changelog_file.exists() or changelog_file.stat().st_size == 0
    now = datetime.now(HST).isoformat(timespec="seconds")
    with changelog_file.open("a", newline="") as f:
        fcntl.flock(f, fcntl.LOCK_EX)
        try:
            writer = csv.writer(f)
            if needs_header:
                writer.writerow(CHANGELOG_HEADER)
            writer.writerow([
                now,
                action,
                event_name,
                ev_date.isoformat(),
                _format_clock(start.time()),
                _format_clock(end.time()),
                details,
            ])
        finally:
            fcntl.flock(f, fcntl.LOCK_UN)


def _format_clock(t: time) -> str:
    hour = t.hour % 12 or 12
    ampm = "AM" if t.hour < 12 else "PM"
    return f"{hour}:{t.minute:02d} {ampm}"


# --- Cancel / Delete -----------------------------------------------------

ACTION_FOR_STATUS = {"cancelled": "CANCEL", "deleted": "DELETE"}


def update_event_status(
    *,
    event_name: str,
    event_date: date,
    new_status: str,
    schedule_dir: Path,
    reminders_file: Path,
    changelog_file: Path,
    details: str | None = None,
) -> int:
    """Cancel or delete the event identified by (event_name, event_date).

    Per scheduler_instructions.md §Cancel vs. Delete:
      - Rewrite every spreadsheet cell containing this active event from
        ``Name`` to ``Name (cancelled)`` / ``Name (deleted)``.
      - Set ``status: new_status`` on every matching reminders.json entry.
      - Append a CANCEL or DELETE row to changelog.csv.

    Returns the number of spreadsheet cells modified.
    """
    if new_status not in _STATUS_SUFFIX:
        raise ValueError(f"Unsupported status: {new_status}")
    name = event_name.strip()
    if not name:
        raise ValueError("event_name is required")

    modified_cells, start_t, end_t = _restatus_cells(
        schedule_dir, name, event_date, new_status
    )
    _restatus_reminders(reminders_file, name, event_date, new_status)
    _append_changelog(
        changelog_file,
        action=ACTION_FOR_STATUS[new_status],
        event_name=name,
        ev_date=event_date,
        start=_attach_hst(event_date, start_t or time(0, 0)),
        end=_attach_hst(event_date, end_t or time(0, 0)),
        details=details or f"{new_status.capitalize()} via mobile app",
    )
    return modified_cells


def _attach_hst(d: date, t: time) -> datetime:
    return datetime.combine(d, t, tzinfo=HST)


def _restatus_cells(
    schedule_dir: Path,
    event_name: str,
    event_date: date,
    new_status: str,
) -> tuple[int, time | None, time | None]:
    """Rewrite cells on event_date where event_name appears as an active entry.

    Returns (cells_modified, earliest_slot, latest_slot+30min).
    """
    ws_start = week_start(event_date)
    year, month_name = workbook_for_week(ws_start)
    wb_path = schedule_dir / str(year) / f"{year}_{month_name}.xlsx"
    if not wb_path.exists():
        return 0, None, None

    sheet_name = sheet_name_for_week(ws_start)
    col = _day_column((event_date - ws_start).days)
    suffix = _STATUS_SUFFIX[new_status]

    cells_modified = 0
    slots_touched: list[time] = []

    with _xlsx_lock(wb_path):
        wb = load_workbook(str(wb_path))
        try:
            if sheet_name not in wb.sheetnames:
                return 0, None, None
            ws = wb[sheet_name]
            for row_idx, slot in enumerate(half_hour_slots(), start=2):
                raw = ws.cell(row=row_idx, column=col).value
                if not raw:
                    continue
                new_text, changed = _annotate_active_match(
                    str(raw), event_name, suffix
                )
                if changed:
                    ws.cell(row=row_idx, column=col, value=new_text)
                    cells_modified += 1
                    slots_touched.append(slot)
            if cells_modified:
                wb.save(str(wb_path))
        finally:
            wb.close()

    if not slots_touched:
        return 0, None, None
    earliest = min(slots_touched)
    latest_plus = _slot_plus_30(max(slots_touched))
    return cells_modified, earliest, latest_plus


def _slot_plus_30(t: time) -> time:
    end_dt = datetime(2000, 1, 1, t.hour, t.minute) + timedelta(minutes=30)
    # Roll midnight back so the "end" still reads as 12:00 AM rather than
    # spilling into the next date; the spreadsheet row range is [0..23:30].
    return end_dt.time()


def _annotate_active_match(
    cell_text: str, event_name: str, suffix: str
) -> tuple[str, bool]:
    """Replace each active ``event_name`` in a cell with ``event_name suffix``.

    Pipe-separated entries already annotated with ``(cancelled)`` or
    ``(deleted)`` are left alone.
    """
    parts = [p.strip() for p in cell_text.split("|")]
    changed = False
    out: list[str] = []
    for part in parts:
        if not part:
            continue
        bare, existing = _strip_status_suffix(part)
        if bare == event_name and existing is None:
            out.append(f"{event_name} {suffix}")
            changed = True
        else:
            out.append(part)
    return " | ".join(out), changed


def _strip_status_suffix(text: str) -> tuple[str, str | None]:
    """Return (bare_name, status) where status is 'cancelled', 'deleted', or None."""
    lower = text.lower()
    for status, suffix in _STATUS_SUFFIX.items():
        if lower.endswith(suffix):
            return text[: -len(suffix)].rstrip(), status
    return text, None


def _restatus_reminders(
    reminders_file: Path, event_name: str, event_date: date, new_status: str
) -> int:
    """Set status on every matching reminder. Returns the count updated."""
    if not reminders_file.exists():
        return 0
    updated = 0
    target_date = event_date.isoformat()
    with reminders_file.open("r+") as f:
        fcntl.flock(f, fcntl.LOCK_EX)
        try:
            try:
                data = json.load(f)
                if not isinstance(data, list):
                    return 0
            except json.JSONDecodeError:
                return 0
            for entry in data:
                if not isinstance(entry, dict):
                    continue
                if (entry.get("event_name") == event_name
                        and entry.get("event_date") == target_date
                        and entry.get("status", "active") == "active"):
                    entry["status"] = new_status
                    updated += 1
            if updated:
                f.seek(0)
                f.truncate()
                json.dump(data, f, indent=2)
        finally:
            fcntl.flock(f, fcntl.LOCK_UN)
    return updated
