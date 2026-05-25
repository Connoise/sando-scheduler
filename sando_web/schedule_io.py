"""Read-side I/O for the master spreadsheet and reminders.json.

Reads are file-locked per `mobile_app_plan.md` §Concurrency: shared lock
on JSON, no lock on xlsx beyond openpyxl's own read_only mode (writes
will be coordinated via a sibling .lock file in v0.3).

Parsed weekly sheets are cached in-process and invalidated when the
workbook's mtime changes.
"""

from __future__ import annotations

import fcntl
import json
import logging
from dataclasses import dataclass
from datetime import date, time, timedelta
from pathlib import Path
from threading import Lock

from openpyxl import load_workbook

from .calendar_utils import (
    WeekContext,
    half_hour_slots,
    parse_slot_label,
    sheet_name_for_week,
    slot_label,
    week_start,
    workbook_for_week,
)


logger = logging.getLogger(__name__)


# Match the spreadsheet's "Event Name (cancelled)" / "(deleted)" annotation.
_STATUS_SUFFIXES = {
    "(cancelled)": "cancelled",
    "(deleted)": "deleted",
}


@dataclass(frozen=True)
class EventCell:
    """One event drawn from a single spreadsheet cell.

    Multiple events in one cell (separated by ' | ' per the spec) become
    multiple EventCell instances sharing the same date+slot.
    """

    date: date
    slot: time
    name: str
    status: str  # "active" | "cancelled" | "deleted"

    @property
    def slot_label(self) -> str:
        return slot_label(self.slot)


def _parse_cell(raw: str) -> list[tuple[str, str]]:
    """Split a cell's text into (name, status) tuples."""
    if not raw:
        return []
    out: list[tuple[str, str]] = []
    for part in str(raw).split("|"):
        text = part.strip()
        if not text:
            continue
        status = "active"
        for suffix, marker in _STATUS_SUFFIXES.items():
            if text.lower().endswith(suffix):
                status = marker
                text = text[: -len(suffix)].rstrip()
                break
        out.append((text, status))
    return out


class WeeklySheetCache:
    """mtime-keyed cache of parsed weekly sheets, per `mobile_app_plan.md`."""

    def __init__(self) -> None:
        self._entries: dict[Path, tuple[float, dict[str, list[EventCell]]]] = {}
        self._lock = Lock()

    def get_week(
        self, workbook_path: Path, ctx: WeekContext
    ) -> list[EventCell]:
        if not workbook_path.exists():
            return []
        mtime = workbook_path.stat().st_mtime
        with self._lock:
            cached = self._entries.get(workbook_path)
            if cached is None or cached[0] != mtime:
                parsed = _parse_workbook(workbook_path)
                self._entries[workbook_path] = (mtime, parsed)
            else:
                parsed = cached[1]
        return parsed.get(ctx.sheet_name, [])

    def clear(self) -> None:
        with self._lock:
            self._entries.clear()


def _week_start_from_sheet_name(sheet_name: str, year: int) -> date | None:
    """Parse a sheet name like 'May 24-30' into the week's start date.

    The spreadsheet uses '<MonthAbbr> <startDay>-<endDay>' per
    scheduler_instructions.md. We only need the month abbreviation and the
    start day; the year comes from the workbook path.
    """
    import calendar as _cal
    parts = sheet_name.strip().split()
    if len(parts) < 2:
        return None
    month_abbr = parts[0]
    day_range = parts[1]
    start_day_str = day_range.split("-")[0]
    try:
        start_day = int(start_day_str)
    except ValueError:
        return None
    abbr_to_num = {abbr: i for i, abbr in enumerate(_cal.month_abbr) if abbr}
    month_num = abbr_to_num.get(month_abbr)
    if month_num is None:
        return None
    try:
        return date(year, month_num, start_day)
    except ValueError:
        return None


def _parse_workbook(path: Path) -> dict[str, list[EventCell]]:
    """Return {sheet_name: [EventCell, ...]} for every weekly sheet."""
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as e:  # openpyxl raises a variety; we want any failure logged
        logger.error("Failed to open workbook %s: %s", path, e)
        return {}

    # Extract year from path (.../2026/2026_May.xlsx → 2026).
    try:
        workbook_year = int(path.parent.name)
    except (ValueError, AttributeError):
        workbook_year = date.today().year

    result: dict[str, list[EventCell]] = {}
    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            week_start_date = _week_start_from_sheet_name(sheet_name, workbook_year)
            result[sheet_name] = list(_parse_sheet(sheet, week_start_date))
    finally:
        wb.close()
    return result


def _parse_sheet(sheet, week_start_date: date | None) -> list[EventCell]:
    """Parse one weekly sheet into EventCell objects.

    Layout (per scheduler_instructions.md §Workbook Layout):
      - Row 1: header. Columns B..H are Sun..Sat.
      - Column A rows 2..49: slot labels like '2:30 PM'.
      - Columns B..H rows 2..49: cell content.

    Dates are derived from week_start_date (preferred) so we don't depend
    on the header row format, which varies across workbooks.
    """
    if week_start_date is not None:
        # Columns B..H (indices 2..8) map to Sun..Sat offset 0..6.
        header_dates: dict[int, date] = {
            col: week_start_date + timedelta(days=col - 2) for col in range(2, 9)
        }
    else:
        # Fallback: try to parse dates from the header row text.
        header_dates = {}
        for col in range(2, 9):
            raw = sheet.cell(row=1, column=col).value
            d = _extract_date_from_header(raw)
            if d is not None:
                header_dates[col] = d

    events: list[EventCell] = []
    for row in range(2, 50):  # 48 slot rows
        label_value = sheet.cell(row=row, column=1).value
        slot = _coerce_slot(label_value)
        if slot is None:
            continue
        for col, ev_date in header_dates.items():
            raw = sheet.cell(row=row, column=col).value
            for name, status in _parse_cell(raw):
                events.append(
                    EventCell(date=ev_date, slot=slot, name=name, status=status)
                )
    return events


def _coerce_slot(value) -> time | None:
    """Time cells may be stored as either a string label or a real time."""
    if isinstance(value, time):
        # openpyxl returns a datetime.time for time-typed cells.
        return value.replace(second=0, microsecond=0)
    if value is None:
        return None
    return parse_slot_label(str(value))


def _extract_date_from_header(value) -> date | None:
    """Header cells look like 'Sunday 4/6' (year inferred from workbook).

    For now we only support the literal month/day form. Year handling
    lives in the caller, which already knows which year's workbook this
    is. We attach the workbook year in `WeeklySheetView.build` below.
    """
    if value is None:
        return None
    if isinstance(value, date):
        return value
    return _parse_header_label(str(value))


def _parse_header_label(label: str) -> date | None:
    """Parse 'Sunday 4/6' style headers. Year is filled in later."""
    label = label.strip()
    if not label:
        return None
    # Take the last whitespace-separated token, which should be 'M/D'.
    token = label.split()[-1]
    if "/" not in token:
        return None
    try:
        month_str, day_str = token.split("/", 1)
        # We don't know the year yet; use a placeholder year 1900.
        return date(1900, int(month_str), int(day_str))
    except ValueError:
        return None


@dataclass(frozen=True)
class Reminder:
    id: str
    event_name: str
    event_date: date | None
    event_start: str
    event_end: str
    status: str

    @classmethod
    def from_dict(cls, raw: dict) -> "Reminder | None":
        if not isinstance(raw, dict):
            return None
        ev_date_str = raw.get("event_date")
        ev_date: date | None = None
        if isinstance(ev_date_str, str):
            try:
                ev_date = date.fromisoformat(ev_date_str)
            except ValueError:
                ev_date = None
        return cls(
            id=str(raw.get("id", "")),
            event_name=str(raw.get("event_name", "")),
            event_date=ev_date,
            event_start=str(raw.get("event_start", "")),
            event_end=str(raw.get("event_end", "")),
            status=str(raw.get("status", "active")),
        )


def load_reminders(path: Path) -> list[Reminder]:
    """Read reminders.json with a shared lock, per mobile_app_plan.md."""
    if not path.exists():
        return []
    try:
        with path.open("r") as f:
            fcntl.flock(f, fcntl.LOCK_SH)
            try:
                data = json.load(f)
            finally:
                fcntl.flock(f, fcntl.LOCK_UN)
    except (OSError, json.JSONDecodeError) as e:
        logger.error("Failed to read %s: %s", path, e)
        return []
    if not isinstance(data, list):
        return []
    out: list[Reminder] = []
    for raw in data:
        r = Reminder.from_dict(raw)
        if r is not None:
            out.append(r)
    return out


# --- View assembly --------------------------------------------------------

@dataclass(frozen=True)
class WeeklySheetView:
    """The full 7×48 grid that the weekly template renders."""

    ctx: WeekContext
    rows: list["SlotRow"]
    workbook_exists: bool

    @classmethod
    def build(
        cls,
        ctx: WeekContext,
        workbook_path: Path,
        cache: WeeklySheetCache,
    ) -> "WeeklySheetView":
        raw_events = cache.get_week(workbook_path, ctx)
        # Fill placeholder years from the workbook's year, then index by
        # (date, slot) for O(1) cell lookup.
        by_cell: dict[tuple[date, time], list[EventCell]] = {}
        for ev in raw_events:
            ev_date = ev.date
            if ev_date.year == 1900:
                ev_date = ev_date.replace(year=ctx.workbook_year)
            by_cell.setdefault((ev_date, ev.slot), []).append(
                EventCell(
                    date=ev_date,
                    slot=ev.slot,
                    name=ev.name,
                    status=ev.status,
                )
            )

        rows: list[SlotRow] = []
        week_dates_ = [ctx.start.fromordinal(ctx.start.toordinal() + i) for i in range(7)]
        for slot in half_hour_slots():
            cells = [by_cell.get((d, slot), []) for d in week_dates_]
            rows.append(SlotRow(slot=slot, label=slot_label(slot), cells=cells))

        return cls(ctx=ctx, rows=rows, workbook_exists=workbook_path.exists())


@dataclass(frozen=True)
class SlotRow:
    slot: time
    label: str
    # Per-day list of EventCells in this slot (length 7, Sunday..Saturday).
    cells: list[list[EventCell]]


# --- Cross-week range queries --------------------------------------------

def load_events_in_range(
    start_date: date,
    end_date: date,
    schedule_dir: Path,
    cache: WeeklySheetCache,
) -> list[EventCell]:
    """Return all EventCells in [start_date, end_date], inclusive.

    Resolves the (workbook, sheet) for every week the range touches and
    pulls events from each, filtering down to the requested dates. The
    cache makes the openpyxl parse cost a one-time hit per workbook
    until its mtime changes.
    """
    from datetime import timedelta as _td

    if end_date < start_date:
        return []

    visited_sheets: set[tuple[Path, str, int]] = set()
    out: list[EventCell] = []

    d = week_start(start_date)
    while d <= end_date:
        year, month_name = workbook_for_week(d)
        workbook_path = schedule_dir / str(year) / f"{year}_{month_name}.xlsx"
        sheet = sheet_name_for_week(d)
        key = (workbook_path, sheet, year)
        if key in visited_sheets:
            d += _td(days=7)
            continue
        visited_sheets.add(key)

        ctx = WeekContext.for_date(d)
        events = cache.get_week(workbook_path, ctx)
        for ev in events:
            ev_date = ev.date
            if ev_date.year == 1900:
                ev_date = ev_date.replace(year=year)
            if start_date <= ev_date <= end_date:
                out.append(
                    EventCell(
                        date=ev_date,
                        slot=ev.slot,
                        name=ev.name,
                        status=ev.status,
                    )
                )
        d += _td(days=7)

    return out


# --- Month view -----------------------------------------------------------

@dataclass(frozen=True)
class MonthDayCell:
    """One day cell on the month grid."""

    date: date
    in_month: bool
    is_today: bool
    events: list[EventCell]

    @property
    def active_count(self) -> int:
        return sum(1 for e in self.events if e.status == "active")

    @property
    def display_events(self) -> list[EventCell]:
        # Show the first two events; the template handles the "+N more" suffix.
        return self.events[:2]

    @property
    def overflow(self) -> int:
        return max(0, len(self.events) - 2)


@dataclass(frozen=True)
class MonthView:
    year: int
    month: int
    month_name: str
    grid_start: date  # Sunday on/before the first of the month
    grid_end: date    # Saturday on/after the last of the month
    weeks: list[list[MonthDayCell]]
    prev_month: tuple[int, int]
    next_month: tuple[int, int]
    today: date

    @classmethod
    def build(
        cls,
        year: int,
        month: int,
        schedule_dir: Path,
        cache: WeeklySheetCache,
        today: date | None = None,
    ) -> "MonthView":
        from calendar import monthrange
        from datetime import timedelta as _td

        from .calendar_utils import MONTH_NAMES

        today = today or date.today()
        first = date(year, month, 1)
        _, days_in_month = monthrange(year, month)
        last = date(year, month, days_in_month)
        grid_start = week_start(first)
        # Walk to the Saturday on/after the last day.
        grid_end_week_start = week_start(last)
        grid_end = grid_end_week_start + _td(days=6)

        events = load_events_in_range(grid_start, grid_end, schedule_dir, cache)
        by_date: dict[date, list[EventCell]] = {}
        for ev in events:
            by_date.setdefault(ev.date, []).append(ev)
        # Sort each day's events by slot time so the preview is stable.
        for items in by_date.values():
            items.sort(key=lambda e: e.slot)

        weeks: list[list[MonthDayCell]] = []
        d = grid_start
        while d <= grid_end:
            row = []
            for _ in range(7):
                row.append(
                    MonthDayCell(
                        date=d,
                        in_month=(d.month == month and d.year == year),
                        is_today=(d == today),
                        events=by_date.get(d, []),
                    )
                )
                d += _td(days=1)
            weeks.append(row)

        prev_month = (year, month - 1) if month > 1 else (year - 1, 12)
        next_month = (year, month + 1) if month < 12 else (year + 1, 1)

        return cls(
            year=year,
            month=month,
            month_name=MONTH_NAMES[month - 1],
            grid_start=grid_start,
            grid_end=grid_end,
            weeks=weeks,
            prev_month=prev_month,
            next_month=next_month,
            today=today,
        )
